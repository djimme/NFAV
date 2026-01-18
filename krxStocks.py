import datetime
import logging
import os

import pandas as pd
import requests
from pandas import DataFrame, ExcelWriter
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment

def getKrxStocks():         
    base_url: str = "https://kind.krx.co.kr/corpgeneral/corpList.do"
    save_path: str = "./corpCode/comp_list.html"

    if not os.path.exists(save_path):            
        """
        KIND 상장종목현황 화면의 EXCEL 버튼과 동일한 요청을 날려 엑셀 파일을 저장한다.
        필요한 경우 params 값을 F12 개발자 도구 내 Network 탭에서 확인 후 수정한다.
        """

        # 아래 params는 예시이며, 실제로는 개발자 도구에서
        # EXCEL 버튼 클릭 시 corpList.do?method=download 로 나가는
        # 쿼리스트링을 그대로 옮겨 적어야 한다.
        params = {
            # 아래부터는 개발자도구에서 보고 그대로 세팅
            "method": "download",   # fnDownload 안에서 지정되는 다운로드용 method 값
            "searchType" : "13",  # 전체검색
        }

        headers = {
            # 최소한의 헤더 (필요 시 개발자 도구에서 복사해서 추가)
            "User-Agent": (
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/143.0.0.0 Safari/537.36 Edg/143.0.0.0"
            ),
            "Referer": "https://kind.krx.co.kr/corpgeneral/corpList.do?method=loadInitPage",
        }

        resp = requests.get(base_url, params=params, headers=headers)
        resp.raise_for_status()

        Path(save_path).write_bytes(resp.content)    
        print(f"saved: {save_path}")

    code_df = pd.read_html(save_path, header=0)[0]      
    
    code_df = code_df[~code_df['회사명'].str.contains('스팩')]
    code_df = code_df[~code_df['시장구분'].str.contains('코넥스')]
    
    code_df.reset_index(inplace=True)  
        
    code_df = code_df[['종목코드', '회사명', '업종', '주요제품']]
    code_df = code_df.rename(columns={'종목코드': 'code', '회사명':'name', '업종' : 'industry', '주요제품' : 'main_product'})

    print("[   Data 예제   ] \n", code_df.head())
    return code_df

def getStocksFnguide():
    """
    FnGuide lookup_data.asp API에서 전체 종목 리스트 추출

    Returns:
        DataFrame: 종목분류, 종목코드, 종목명, 시장정보 컬럼을 가진 DataFrame
    """
    url = "https://comp.fnguide.com/SVO2/common/lookup_data.asp"
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
        'Referer': 'https://comp.fnguide.com/SVO2/common/lookup.asp'
    }

    # 종목 분류 목록 (cmbCompGb의 value 값)
    # 0: 전체, 1: Company, 2: ETF, 8: ETN, 3: REITs, 4: SPAC
    comp_type_map = {
        '1': 'Company',
        '2': 'ETF',
        '8': 'ETN',
        '3': 'REITs',
        '4': 'SPAC'
    }

    all_stocks = []

    for comp_gb, comp_name in comp_type_map.items():
        print(f"[FnGuide] {comp_name} 종목 수집 중...")

        # lookup_data.asp API 파라미터
        # mkt_gb: 1(전체), 2(KOSPI), 3(KOSDAQ), 7(KONEX), 8(K-OTC), 6(상장예정)
        # comp_gb: 종목분류
        # s_type: 1(일반검색), 2(가나다 검색)
        # search_key1, search_key2: 검색어
        params = {
            'mkt_gb': '1',  # 전체 시장
            'comp_gb': comp_gb,
            's_type': '1',  # 일반 검색
            'search_key1': '',  # 전체
            'search_key2': ''
        }

        try:
            response = requests.get(url, params=params, headers=headers, timeout=30)
            response.encoding = 'utf-8'

            # JSON 응답 파싱
            data = response.json()

            # 데이터 처리
            # data는 [{cd: 'A005930', nm: '삼성전자', gb: '코스피', stk_gb: '701', mkt_gb: '2'}, ...] 형식
            for item in data:
                # 종목코드에서 'A' 제거
                stock_code = item.get('cd', '')
                if stock_code.startswith(('A', 'Q')) and len(stock_code) == 7:
                    stock_code = stock_code[1:]  # A 또는 Q 제거
                else:
                    continue

                stock_name = item.get('nm', '')
                market_info = item.get('gb', '')

                all_stocks.append({
                    '종목분류': comp_name,
                    '종목코드': stock_code,
                    '종목명': stock_name,
                    '시장정보': market_info
                })

            print(f"  {comp_name}: {len([s for s in all_stocks if s['종목분류'] == comp_name])}개 종목")

        except Exception as e:
            print(f"  ERROR - {comp_name} 수집 실패: {e}")
            continue

    # DataFrame 생성
    if all_stocks:
        df = pd.DataFrame(all_stocks)
        print(f"\n[FnGuide] 총 {len(df)}개 종목 수집 완료")
        print(f"[   Data 예제   ] \n{df.head()}")
        return df
    else:
        print("\n[FnGuide] 수집된 종목이 없습니다.")
        return pd.DataFrame(columns=['종목분류', '종목코드', '종목명', '시장정보'])


if __name__ == '__main__':
    now = datetime.datetime.now()
    corp_list:str = "./corpCode/corplist_{0}{1:02d}.xlsx".format(now.year, now.month)
    flag = 0

    if not os.path.exists(corp_list):
        print("="*10 + " Get KRX Stocks " + "="*10)
        if flag == 1:
            df_complist = getKrxStocks()
        else:
            df_complist = getStocksFnguide()

        df_complist.to_excel(corp_list, sheet_name="CorpList", index=False)

        # 엑셀 서식 적용
        wb = load_workbook(corp_list)
        ws = wb['CorpList']

        # 폰트 및 정렬 스타일 정의
        font_style = Font(name='맑은 고딕', size=10)
        alignment_style = Alignment(horizontal='center', vertical='center')

        # 모든 셀에 서식 적용
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.font = font_style
                cell.alignment = alignment_style

        wb.save(corp_list)
        print(f"엑셀 파일 저장 완료: {corp_list}")