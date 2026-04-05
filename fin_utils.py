"""
FnGuide 스크래핑 공통 유틸리티

Excel 저장:
  save_styled_excel(df, filepath)             : 단일 시트 서식 저장
  save_styled_excel_multisheet(sheets, filepath): 멀티시트 서식 저장

HTTP / 캐싱:
  fetch_fnguide_page(code, asp_page, menu_id, cache_prefix): FnGuide 페이지 다운로드 (캐싱)

HTML 파싱 공통 헬퍼 (모든 FnGuide 모듈에서 공유):
  parse_company_name(soup): 페이지 title에서 종목명 추출
  parse_kse_fics(soup)    : KSE/FICS 분야 및 결산월 추출
"""
import datetime
import os
import re
import shutil

import requests
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows


def save_styled_excel(df, filepath, sheet_name="Sheet1", index=False):
    """DataFrame을 서식이 적용된 Excel 파일로 저장한다.

    서식:
        - 글꼴: 맑은 고딕, 10pt
        - 맞춤: 상하 가운데, 좌우 가운데
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    font = Font(name="맑은 고딕", size=10)
    alignment = Alignment(horizontal="center", vertical="center")

    for row in dataframe_to_rows(df, index=index, header=True):
        ws.append(row)

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.font = font
            cell.alignment = alignment
            if isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0.##'

    wb.save(filepath)


def save_styled_excel_multisheet(sheets, filepath):
    """여러 시트를 담은 서식 적용 Excel 파일 저장

    Args:
        sheets  : [(sheet_name, DataFrame), ...] 순서 있는 시트 목록
        filepath: 저장 경로 (예: 'derived/output.xlsx')

    서식:
        - 글꼴: 맑은 고딕, 10pt
        - 맞춤: 상하 가운데, 좌우 가운데
    """
    wb = Workbook()
    font = Font(name="맑은 고딕", size=10)
    alignment = Alignment(horizontal="center", vertical="center")

    first = True
    for sheet_name, df in sheets:
        if first:
            ws = wb.active
            ws.title = sheet_name
            first = False
        else:
            ws = wb.create_sheet(title=sheet_name)

        for row in dataframe_to_rows(df, index=False, header=True):
            ws.append(row)

        for row_cells in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
            for cell in row_cells:
                cell.font = font
                cell.alignment = alignment
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0.##'

    wb.save(filepath)


def fetch_fnguide_page(code, asp_page, menu_id, cache_prefix):
    """FnGuide 페이지를 다운로드하고 캐싱하는 공통 함수

    Parameters:
        code         : 종목코드 (6자리 문자열)
        asp_page     : FnGuide ASP 페이지명 (예: 'SVD_Finance.asp')
        menu_id      : FnGuide NewMenuID 값 (예: '103')
        cache_prefix : 캐시 디렉토리 접두어 (예: 'fnguide_finance_')

    Returns:
        str : HTML 문자열
    """
    now = datetime.datetime.now()
    url = f"https://comp.fnguide.com/SVO2/ASP/{asp_page}?pGB=1&gicode=A{code}&cID=&MenuYn=Y&ReportGB=&NewMenuID={menu_id}&stkGb=701"

    downloadedFileDirectory = 'derived/{0}{1}-{2:02d}'.format(cache_prefix, now.year, now.month)
    downloadedFilePath = '{0}/{1}.html'.format(downloadedFileDirectory, code)

    os.makedirs(downloadedFileDirectory, exist_ok=True)

    # 이전 달 캐시 디렉토리 삭제
    currentDirName = os.path.basename(downloadedFileDirectory)
    for entry in os.listdir('derived'):
        if entry.startswith(cache_prefix) and entry != currentDirName:
            shutil.rmtree(os.path.join('derived', entry))

    if os.path.exists(downloadedFilePath):
        return open(downloadedFilePath, "r", encoding="utf-8").read()

    response = requests.get(url)
    response.encoding = 'utf-8'
    with open(downloadedFilePath, "w", encoding="utf-8") as f:
        f.write(response.text)
    return response.text


# ── FnGuide HTML 공통 파싱 헬퍼 ────────────────────────────────────────────────
# 모든 FnGuide 페이지(Snapshot/Finance/FinanceRatio/InvestIdx)가 공유하는
# 공통 헤더 파싱 로직. 각 모듈에 중복 선언하지 않고 여기서 import하여 사용한다.

def parse_company_name(soup) -> str:
    """페이지 title 태그에서 종목명 추출

    FnGuide 페이지 title 형식: '삼성전자(A005930) ...' → '삼성전자'
    """
    title = soup.find('title')
    if title:
        match = re.match(r'^([^(]+)\(A\d+\)', title.get_text(strip=True))
        if match:
            return match.group(1).strip()
    return ''


def parse_kse_fics(soup) -> tuple:
    """KSE/FICS 분야 및 결산월 추출 → (kse_sector, fics_sector, fiscal_month)

    - KSE/FICS : p.stxt_group 내 span.stxt
    - 결산월    : div.corp_group1 > h2 에서 "12월 결산" 형식 (없으면 None)
    """
    kse_sector = ''
    fics_sector = ''
    fiscal_month = None

    corp_group1 = soup.find('div', class_='corp_group1')
    if corp_group1:
        for h2 in corp_group1.find_all('h2'):
            m = re.search(r'(\d+)월\s*결산', h2.get_text(strip=True))
            if m:
                fiscal_month = int(m.group(1))
                break

    stxt_group = soup.find('p', class_='stxt_group')
    if not stxt_group:
        return kse_sector, fics_sector, fiscal_month

    for span in stxt_group.find_all('span', class_='stxt'):
        text = span.get_text(strip=True).replace('\xa0', ' ')
        if text.startswith('FICS'):
            fics_sector = re.sub(r'^FICS\s+', '', text).strip()
        else:
            for prefix in ('KSE', 'KOSDAQ', 'K-OTC', 'KONEX'):
                if text.startswith(prefix):
                    kse_sector = re.sub(rf'^{prefix}\s+', '', text).strip()
                    break

    return kse_sector, fics_sector, fiscal_month
